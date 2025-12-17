
import re
from io import BytesIO
from datetime import datetime
import textwrap

import pandas as pd
import streamlit as st
from docx import Document
from openpyxl import Workbook

from PIL import Image
from io import BytesIO


from streamlit_drawable_canvas import st_canvas


# Optional PDF dependency
REPORTLAB_AVAILABLE = True
try:
    from reportlab.lib.pagesizes import A4
    from reportlab.pdfgen import canvas
    from reportlab.lib.units import mm
except Exception:
    REPORTLAB_AVAILABLE = False

# ============================
# CONFIG (Stage 2 coding)
# ============================
EMPATHY_QUADRANTS = [
    "thinks_feels",
    "how_where_when_tools",
    "motivation",
    "what_to_learn",
]

PAIN_GAIN = ["neutral", "pain", "gain"]
CODING_CONFIDENCE = ["high", "medium", "low"]

EMOTIONS = [
    "", "stress", "anxiety", "frustration", "overwhelm", "confusion",
    "confidence", "relief", "interest", "motivation", "pride"
]

LEARNING_FOCUS = [
    "", "content", "strategy", "tool_use", "time_management", "assessment",
    "social_support", "self_efficacy", "attention_distraction", "wellbeing"
]

# ============================
# Word parsing (Stage 1)
# ============================
QUESTION_RE = re.compile(r'^([A-G]\d+)\.\s')  # e.g., "A1. ..."

SECTION_MAP = {
    "A": "Study Context",
    "B": "Learning Content & Structure",
    "C": "Learning Practices",
    "D": "Learning Environment",
    "E": "Motivation",
    "F": "Challenging Experiences (Pain)",
    "G": "Positive Experiences (Gain)",
}

DEFAULT_PAIN_GAIN_BY_SECTION = {"F": "pain", "G": "gain"}

DEFAULT_EMPATHY_BY_SECTION = {
    "A": "thinks_feels",  # NOTE: you may want to set A1 to what_to_learn manually
    "B": "what_to_learn",
    "C": "how_where_when_tools",
    "D": "how_where_when_tools",
    "E": "motivation",
    "F": "thinks_feels",
    "G": "thinks_feels",
}

def clean_text(s: str) -> str:
    if not s:
        return ""
    s = s.replace("\u00A0", " ")
    s = re.sub(r'[ \t]+', ' ', s)
    s = re.sub(r'\n{3,}', '\n\n', s)
    return s.strip()

def split_statements(answer: str):
    if not answer:
        return []
    lines = []
    for raw in answer.splitlines():
        line = raw.strip()
        if not line:
            continue
        line = re.sub(r'^[\u2022\-\*\u00B7]+\s*', '', line).strip()
        if line:
            lines.append(line)
    if len(lines) == 1 and ";" in lines[0]:
        parts = [p.strip() for p in lines[0].split(";") if p.strip()]
        if len(parts) > 1:
            return parts
    return lines

def extract_inline_answer(paragraph_text: str) -> str:
    if "ANSWER:" not in paragraph_text:
        return ""
    after = paragraph_text.split("ANSWER:", 1)[1]
    return clean_text(after)

def extract_answer_after(paras, idx, stop_pred):
    text = paras[idx].strip()
    collected = []
    if text.startswith("ANSWER:"):
        after = text[len("ANSWER:"):].strip()
        if after:
            collected.append(after)
    j = idx + 1
    while j < len(paras):
        t = paras[j].strip()
        if stop_pred(t):
            break
        if t.startswith("ANSWER:"):
            break
        if t:
            collected.append(t)
        j += 1
    return clean_text("\n".join(collected)), j

def parse_docx(file_bytes: bytes):
    doc = Document(BytesIO(file_bytes))
    paras = [p.text if p.text is not None else "" for p in doc.paragraphs]

    meta = {"participant_id": "", "interview_id": "", "study_phase": ""}

    for t in [p.strip() for p in paras]:
        if t.startswith("Participant ID"):
            meta["participant_id"] = extract_inline_answer(t)
        if t.startswith("Interview ID"):
            meta["interview_id"] = extract_inline_answer(t)
        if t.startswith("Study phase"):
            meta["study_phase"] = extract_inline_answer(t)

    raw_rows = []
    i = 0
    while i < len(paras):
        t = paras[i].strip()
        m = QUESTION_RE.match(t)
        if not m:
            i += 1
            continue

        qid = m.group(1)
        section_letter = qid[0]
        section_name = SECTION_MAP.get(section_letter, "")

        j = i + 1
        while j < len(paras) and not paras[j].strip().startswith("ANSWER:"):
            if QUESTION_RE.match(paras[j].strip()):
                break
            j += 1

        answer = ""
        nxt = i + 1
        if j < len(paras) and paras[j].strip().startswith("ANSWER:"):
            answer, nxt = extract_answer_after(
                paras, j,
                lambda x: bool(QUESTION_RE.match(x)) or x.startswith(("A.", "B.", "C.", "D.", "E.", "F.", "G."))
            )

        for stmnt in split_statements(answer):
            raw_rows.append({
                "participant_id": meta.get("participant_id",""),
                "interview_id": meta.get("interview_id",""),
                "question_id": qid,
                "question_section": section_name,
                "quote": stmnt,
                "study_phase": meta.get("study_phase",""),
            })

        i = max(nxt, i + 1)

    return meta, pd.DataFrame(raw_rows)

# ============================
# Stage 2 builder
# ============================
def build_empathy_coding_df(raw_df: pd.DataFrame):
    if raw_df.empty:
        cols = [
            "participant_id","interview_id","quote_id","question_id","quote",
            "empathy_quadrant_primary","empathy_quadrant_secondary",
            "pain_gain","emotion","learning_focus","coding_confidence","analytic_notes",
            "study_phase"
        ]
        return pd.DataFrame(columns=cols)

    df = raw_df.copy().reset_index(drop=True)
    if "quote_id" not in df.columns or df["quote_id"].astype(str).str.strip().eq("").all():
        df.insert(2, "quote_id", [f"Q{str(i+1).zfill(4)}" for i in range(len(df))])

    def default_empathy(row):
        sec = str(row.get("question_id",""))[:1]
        return DEFAULT_EMPATHY_BY_SECTION.get(sec, "")

    def default_pain_gain(row):
        sec = str(row.get("question_id",""))[:1]
        return DEFAULT_PAIN_GAIN_BY_SECTION.get(sec, "neutral")

    out = pd.DataFrame({
        "participant_id": df["participant_id"].astype(str),
        "interview_id": df["interview_id"].astype(str),
        "quote_id": df["quote_id"].astype(str),
        "question_id": df["question_id"].astype(str),
        "quote": df["quote"].astype(str),

        "empathy_quadrant_primary": df.apply(default_empathy, axis=1),
        "empathy_quadrant_secondary": "",

        "pain_gain": df.apply(default_pain_gain, axis=1),
        "emotion": "",
        "learning_focus": "",
        "coding_confidence": "high",
        "analytic_notes": "",

        "study_phase": df.get("study_phase","").astype(str),
    })
    return out

def df_to_csv_bytes(df: pd.DataFrame) -> bytes:
    return df.to_csv(index=False, quoting=1).encode("utf-8")  # QUOTE_ALL

def to_excel_bytes(meta_list, raw_df: pd.DataFrame, empathy_df: pd.DataFrame, summary_df: pd.DataFrame | None = None):
    wb = Workbook()

    ws_meta = wb.active
    ws_meta.title = "interview_metadata"
    ws_meta.append(["participant_id", "interview_id", "study_phase", "notes"])
    for meta in meta_list:
        ws_meta.append([meta.get("participant_id",""), meta.get("interview_id",""), meta.get("study_phase",""), ""])

    ws_raw = wb.create_sheet("raw_quotes")
    raw_cols = ["participant_id","interview_id","question_id","question_section","quote","study_phase"]
    ws_raw.append(raw_cols)
    if not raw_df.empty:
        for row in raw_df[raw_cols].itertuples(index=False):
            ws_raw.append(list(row))

    ws_emp = wb.create_sheet("empathy_coding")
    emp_cols = list(empathy_df.columns)
    ws_emp.append(emp_cols)
    if not empathy_df.empty:
        for row in empathy_df[emp_cols].itertuples(index=False):
            ws_emp.append(list(row))

    if summary_df is not None:
        ws_sum = wb.create_sheet("empathy_summary")
        sum_cols = list(summary_df.columns)
        ws_sum.append(sum_cols)
        for row in summary_df[sum_cols].itertuples(index=False):
            ws_sum.append(list(row))

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio.getvalue()

# ============================
# Stage 3 synthesis
# ============================
QUADRANTS = [
    ("thinks_feels", "THINKS & FEELS (12)"),
    ("how_where_when_tools", "HOW / WHERE / WHEN / TOOLS (3)"),
    ("motivation", "MOTIVATION (6)"),
    ("what_to_learn", "WHAT TO LEARN (9)"),
]

def build_summary(empathy_df: pd.DataFrame, max_quotes_per_bucket=None) -> pd.DataFrame:
    rows = []
    df = empathy_df.copy()
    for pid, g in df.groupby("participant_id", sort=True):
        def collect(mask):
            sub = g[mask].copy()
            if max_quotes_per_bucket is not None and len(sub) > max_quotes_per_bucket:
                sub = sub.head(max_quotes_per_bucket)
            quotes = sub["quote"].tolist()
            if not quotes:
                return ""
            return "\n".join([f"• {q}" for q in quotes])

        quad_text = {}
        for key, _label in QUADRANTS:
            mask = (g["empathy_quadrant_primary"] == key) | (g["empathy_quadrant_secondary"] == key)
            quad_text[key] = collect(mask)

        pains = collect(g["pain_gain"] == "pain")
        gains = collect(g["pain_gain"] == "gain")

        interview_id = next((x for x in g["interview_id"].tolist() if x.strip()), "")
        study_phase = next((x for x in g["study_phase"].tolist() if x.strip()), "")

        rows.append({
            "participant_id": pid,
            "interview_id": interview_id,
            "study_phase": study_phase,
            "thinks_feels": quad_text["thinks_feels"],
            "how_where_when_tools": quad_text["how_where_when_tools"],
            "motivation": quad_text["motivation"],
            "what_to_learn": quad_text["what_to_learn"],
            "pains": pains,
            "gains": gains,
        })
    return pd.DataFrame(rows)

def wrap_text(text, width_chars=95):
    if not text:
        return []
    out = []
    for line in text.splitlines():
        if not line.strip():
            out.append("")
            continue
        prefix = ""
        body = line
        if line.lstrip().startswith("•"):
            idx = line.find("•")
            prefix = line[:idx+1] + " "
            body = line[idx+1:].strip()
        wrapped = textwrap.wrap(body, width=width_chars)
        if not wrapped:
            out.append(prefix.strip())
        else:
            out.append(prefix + wrapped[0])
            for w in wrapped[1:]:
                out.append("  " + w)
    return out

def draw_box(c, x, y, w, h, title, body_lines, font_size=9):
    c.rect(x, y, w, h, stroke=1, fill=0)
    c.setFont("Helvetica-Bold", 10)
    c.drawString(x + 3*mm, y + h - 6*mm, title)
    c.setFont("Helvetica", font_size)

    top = y + h - 10*mm
    line_h = 4.2*mm
    max_lines = int((h - 14*mm) / line_h)
    lines = body_lines[:max_lines]
    yy = top
    for ln in lines:
        c.drawString(x + 3*mm, yy, ln[:140])
        yy -= line_h

    if len(body_lines) > max_lines:
        c.setFont("Helvetica-Oblique", 8)
        c.drawString(x + 3*mm, y + 3*mm, f"... ({len(body_lines)-max_lines} more lines not shown)")

def draw_classic_empathy_pdf_page(c, page_w, page_h, row):
    # Layout
    margin = 12 * mm
    gap = 6 * mm

    # Top big rectangle + bottom boxes (match canvas style)
    top_h = (page_h - 2*margin - 2*gap) * 0.68
    bottom_h = (page_h - 2*margin - gap) - top_h

    x0 = margin
    y0 = page_h - margin - 24*mm - top_h  # leave room for title area
    w = page_w - 2*margin
    h = top_h

    yb = y0 - gap - bottom_h
    box_w = (w - gap) / 2

    # Title
    pid = row.get("participant_id", "")
    iid = row.get("interview_id", "")
    phase = row.get("study_phase", "")

    c.setFont("Helvetica-Bold", 14)
    c.drawString(margin, page_h - margin, f"Empathy Map (Learning) — {pid}")
    c.setFont("Helvetica", 10)
    c.drawString(margin, page_h - margin - 6*mm, f"Interview: {iid}   Study phase: {phase}")

    # --- Draw main rectangle ---
    c.rect(x0, y0, w, h, stroke=1, fill=0)

    # Diagonals (X)
    c.setLineWidth(1.2)
    c.line(x0, y0, x0 + w, y0 + h)
    c.line(x0 + w, y0, x0, y0 + h)

    # Bottom boxes: Pains / Gains
    c.rect(x0, yb, box_w, bottom_h, stroke=1, fill=0)
    c.rect(x0 + box_w + gap, yb, box_w, bottom_h, stroke=1, fill=0)

    # Labels (positions similar to canvas)
    c.setFont("Helvetica-Bold", 10)
    c.drawString(x0 + w*0.40, y0 + h*0.92, "THINKS & FEELS (12)")
    c.drawString(x0 + w*0.68, y0 + h*0.52, "HOW / WHERE / WHEN / TOOLS (3)")
    c.drawString(x0 + w*0.45, y0 + h*0.10, "MOTIVATION (6)")
    c.drawString(x0 + w*0.05, y0 + h*0.52, "WHAT TO LEARN (9)")

    c.drawString(x0 + 3*mm, yb + bottom_h - 6*mm, "PAINS")
    c.drawString(x0 + box_w + gap + 3*mm, yb + bottom_h - 6*mm, "GAINS")

    # --- Helper to print wrapped bullet text inside a rectangle region ---
    def draw_wrapped_in_region(x, y, w, h, text, font_size=9, top_pad_mm=10):
        c.setFont("Helvetica", font_size)
        lines = wrap_text(text, width_chars=90)  # you already have wrap_text()
        line_h = 4.2 * mm
        max_lines = int((h - top_pad_mm*mm) / line_h)
        lines = lines[:max_lines]

        yy = y + h - top_pad_mm*mm
        for ln in lines:
            c.drawString(x + 3*mm, yy, ln[:140])
            yy -= line_h

   # --- Place content (refined alignment) ---

    # 12 o'clock – THINKS & FEELS (centered)
    draw_wrapped_in_region(
        x0 + w*0.30,          # centered horizontally
        y0 + h*0.55,
        w*0.40,               # narrower width → visual centering
        h*0.40,
        row.get("thinks_feels", "")
    )

    # 3 o'clock – HOW / WHERE / WHEN / TOOLS (closer to right border)
    draw_wrapped_in_region(
        x0 + w*0.67,          # shifted right
        y0 + h*0.25,
        w*0.35,
        h*0.50,
        row.get("how_where_when_tools", "")
    )

    # 6 o'clock – MOTIVATION (centered)
    draw_wrapped_in_region(
        x0 + w*0.35,          # centered horizontally
        y0 + h*0.02,
        w*0.40,
        h*0.35,
        row.get("motivation", "")
    )

    # 9 o'clock – WHAT TO LEARN (closer to left border)
    draw_wrapped_in_region(
        x0 + w*0.03,          # close to left edge
        y0 + h*0.25,
        w*0.35,
        h*0.50,
        row.get("what_to_learn", "")
    )


    # Pains / Gains
    draw_wrapped_in_region(
        x0, yb, box_w, bottom_h,
        row.get("pains", ""),
        top_pad_mm=12
    )
    draw_wrapped_in_region(
        x0 + box_w + gap, yb, box_w, bottom_h,
        row.get("gains", ""),
        top_pad_mm=12
    )


def build_empathy_pdf(summary_df: pd.DataFrame) -> bytes:
    if not REPORTLAB_AVAILABLE:
        raise RuntimeError("reportlab is not installed")

    bio = BytesIO()
    c = canvas.Canvas(bio, pagesize=A4)
    page_w, page_h = A4

    for _, row in summary_df.iterrows():
        draw_classic_empathy_pdf_page(c, page_w, page_h, row)
        c.showPage()

    c.save()
    bio.seek(0)
    return bio.getvalue()


# ============================
# UI
# ============================
st.set_page_config(page_title="Empathy Pipeline (All-in-One)", layout="centered")
st.title("Empathy Interview Pipeline (All-in-One)")
st.caption("Single interface for Stage 1 (Word → raw quotes), Stage 2 (coding with dropdowns), and Stage 3 (empathy map synthesis to summary CSV + optional PDF).")

if not REPORTLAB_AVAILABLE:
    st.info("PDF export is disabled because `reportlab` is not installed in your environment. Install it with: `pip install reportlab` (then restart Streamlit).")

# Keep data across tabs
if "meta_list" not in st.session_state:
    st.session_state.meta_list = []
if "raw_df" not in st.session_state:
    st.session_state.raw_df = pd.DataFrame()
if "empathy_df" not in st.session_state:
    st.session_state.empathy_df = pd.DataFrame()
if "summary_df" not in st.session_state:
    st.session_state.summary_df = pd.DataFrame()

tab1, tab2, tab3, tab4 = st.tabs(["Stage 1: Word → Raw", "Stage 2: Coding", "Stage 3: Synthesis", "Stage 4: Canvas Map"])

def empathy_canvas_template(width=1200, height=850):
    # Layout parameters
    margin = 30
    gap = 20

    # Top big rectangle area + bottom pain/gain boxes
    top_h = int((height - 2*margin - 2*gap) * 0.68)
    bottom_h = height - 2*margin - gap - top_h

    x0 = margin
    y0 = margin
    w = width - 2*margin
    h = top_h

    yb = y0 + h + gap
    box_w = int((w - gap) / 2)

    def rect(x, y, w, h, stroke_width=3):
        return {
            "type": "rect",
            "left": x,
            "top": y,
            "width": w,
            "height": h,
            "fill": "rgba(0,0,0,0)",
            "stroke": "black",
            "strokeWidth": stroke_width,
            "selectable": False,
            "evented": False,
        }

    def line(x1, y1, x2, y2, stroke_width=3):
        return {
            "type": "line",
            "x1": x1, "y1": y1,
            "x2": x2, "y2": y2,
            "stroke": "black",
            "strokeWidth": stroke_width,
            "selectable": False,
            "evented": False,
        }

    def label(x, y, s, size=18):
        return {
            "type": "text",
            "left": x,
            "top": y,
            "text": s,
            "fontSize": size,
            "fontFamily": "Arial",
            "fill": "black",
            "selectable": False,
            "evented": False,
        }

    objects = []

    # --- Top empathy rectangle ---
    objects.append(rect(x0, y0, w, h))

    # Diagonals (X)
    objects.append(line(x0, y0, x0 + w, y0 + h))
    objects.append(line(x0 + w, y0, x0, y0 + h))

    # --- Bottom pain/gain boxes ---
    objects.append(rect(x0, yb, box_w, bottom_h))
    objects.append(rect(x0 + box_w + gap, yb, box_w, bottom_h))

    # --- Labels in the four triangle areas (approx positions) ---
    # Top triangle
    objects.append(label(x0 + w*0.42, y0 + h*0.08, "THINKS & FEELS (12)"))
    # Right triangle
    objects.append(label(x0 + w*0.70, y0 + h*0.42, "HOW / WHERE / WHEN / TOOLS (3)", size=16))
    # Bottom triangle
    objects.append(label(x0 + w*0.46, y0 + h*0.78, "MOTIVATION (6)"))
    # Left triangle
    objects.append(label(x0 + w*0.05, y0 + h*0.42, "WHAT TO LEARN (9)", size=16))

    # Bottom labels
    objects.append(label(x0 + 10, yb + 10, "PAINS", size=18))
    objects.append(label(x0 + box_w + gap + 10, yb + 10, "GAINS", size=18))

    return {"version": "5.2.4", "objects": objects}


# -------- Stage 1 --------
with tab1:
    st.subheader("Stage 1: Upload Word interview sheets and extract quotes")
    uploaded_files = st.file_uploader(
        "Upload one or more .docx files",
        type=["docx"],
        accept_multiple_files=True,
        key="stage1_uploader"
    )

    if uploaded_files:
        all_meta = []
        all_raw = []
        errors = []

        seen_participants = set()
        seen_interviews = set()
        p_counter = 1
        i_counter = 1

        for f in uploaded_files:
            try:
                meta, raw = parse_docx(f.getvalue())

                # ---- Fallback IDs if missing / duplicated ----
                pid = (meta.get("participant_id") or "").strip()
                iid = (meta.get("interview_id") or "").strip()

                if not pid:
                    pid = f"P{p_counter:02d}"
                    p_counter += 1

                if pid in seen_participants:
                    k = 2
                    new_pid = f"{pid}_dup{k}"
                    while new_pid in seen_participants:
                        k += 1
                        new_pid = f"{pid}_dup{k}"
                    pid = new_pid

                if not iid:
                    iid = f"I{i_counter:02d}"
                    i_counter += 1

                if iid in seen_interviews:
                    k = 2
                    new_iid = f"{iid}_dup{k}"
                    while new_iid in seen_interviews:
                        k += 1
                        new_iid = f"{iid}_dup{k}"
                    iid = new_iid

                seen_participants.add(pid)
                seen_interviews.add(iid)

                # Apply resolved IDs back to meta + all extracted rows from this file
                meta["participant_id"] = pid
                meta["interview_id"] = iid

                raw = raw.copy()
                raw["participant_id"] = pid
                raw["interview_id"] = iid

                raw.insert(0, "source_file", f.name)
                all_meta.append({**meta, "source_file": f.name})
                all_raw.append(raw)

            except Exception as e:
                errors.append((f.name, str(e)))

        if errors:
            st.error("Some files could not be parsed:")
            for name, err in errors:
                st.write(f"- {name}: {err}")

        combined_raw = pd.concat(all_raw, ignore_index=True) if all_raw else pd.DataFrame()

        st.session_state.meta_list = [{k: v for k, v in m.items() if k != "source_file"} for m in all_meta]
        st.session_state.raw_df = combined_raw.drop(columns=["source_file"], errors="ignore")

        st.write(f"Rows extracted: **{len(combined_raw)}**")

        # show which participants were detected (helps debugging)
        if not combined_raw.empty and "participant_id" in combined_raw.columns:
            st.info("Participants detected: " + ", ".join(sorted(combined_raw["participant_id"].unique().tolist())))

        st.dataframe(combined_raw, use_container_width=True, hide_index=True)

        st.success("Stage 1 complete. Go to Stage 2 to code the extracted quotes.")
    else:
        st.info("Upload Word files to start the pipeline.")

# -------- Stage 2 --------
with tab2:
    st.subheader("Stage 2: Code quotes for the learning-adapted empathy map (dropdowns)")
    if st.session_state.raw_df is None or st.session_state.raw_df.empty:
        st.warning("No raw quotes available yet. Please complete Stage 1 first.")
    else:
        if st.session_state.empathy_df is None or st.session_state.empathy_df.empty:
            st.session_state.empathy_df = build_empathy_coding_df(st.session_state.raw_df)

        st.caption("Edit the coding columns. IDs and quotes are locked to preserve traceability.")
        edited = st.data_editor(
            st.session_state.empathy_df,
            use_container_width=True,
            hide_index=True,
            num_rows="fixed",
            column_config={
                "empathy_quadrant_primary": st.column_config.SelectboxColumn(
                    "empathy_quadrant_primary", options=EMPATHY_QUADRANTS, required=True
                ),
                "empathy_quadrant_secondary": st.column_config.SelectboxColumn(
                    "empathy_quadrant_secondary", options=[""] + EMPATHY_QUADRANTS, required=False
                ),
                "pain_gain": st.column_config.SelectboxColumn(
                    "pain_gain", options=PAIN_GAIN, required=True
                ),
                "emotion": st.column_config.SelectboxColumn(
                    "emotion", options=EMOTIONS, required=False
                ),
                "learning_focus": st.column_config.SelectboxColumn(
                    "learning_focus", options=LEARNING_FOCUS, required=False
                ),
                "coding_confidence": st.column_config.SelectboxColumn(
                    "coding_confidence", options=CODING_CONFIDENCE, required=True
                ),
                "analytic_notes": st.column_config.TextColumn("analytic_notes"),
                "quote": st.column_config.TextColumn("quote", disabled=True),
                "quote_id": st.column_config.TextColumn("quote_id", disabled=True),
                "participant_id": st.column_config.TextColumn("participant_id", disabled=True),
                "interview_id": st.column_config.TextColumn("interview_id", disabled=True),
                "question_id": st.column_config.TextColumn("question_id", disabled=True),
                "study_phase": st.column_config.TextColumn("study_phase", disabled=True),
            },
        )

        st.session_state.empathy_df = edited

        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        st.download_button(
            "Download Stage-2 empathy coding CSV",
            data=df_to_csv_bytes(edited),
            file_name=f"empathy_coding_{ts}.csv",
            mime="text/csv",
        )

        st.success("Stage 2 ready. Go to Stage 3 to synthesize empathy maps.")

# -------- Stage 3 --------
with tab3:
    st.subheader("Stage 3: Synthesize empathy maps (per participant)")
    if st.session_state.empathy_df is None or st.session_state.empathy_df.empty:
        st.warning("No coded empathy data available yet. Please complete Stage 2 first.")
    else:
        col1, col2 = st.columns(2)
        with col1:
            max_quotes = st.number_input("Max quotes per bucket (0 = no limit)", min_value=0, value=0, step=1)
        with col2:
            make_pdf = st.checkbox("Generate PDF (A4, one page per participant)", value=REPORTLAB_AVAILABLE, disabled=not REPORTLAB_AVAILABLE)

        max_q = None if int(max_quotes) == 0 else int(max_quotes)
        summary = build_summary(st.session_state.empathy_df, max_quotes_per_bucket=max_q)
        st.session_state.summary_df = summary

        st.write(f"Participants detected: **{len(summary)}**")
        st.dataframe(summary, use_container_width=True, hide_index=True)

        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        st.download_button(
            "Download empathy map summary CSV",
            data=df_to_csv_bytes(summary),
            file_name=f"empathy_maps_{ts}_summary.csv",
            mime="text/csv",
        )

        # Combined Excel (meta + raw + coding + summary)
        excel_bytes = to_excel_bytes(
            st.session_state.meta_list if st.session_state.meta_list else [],
            st.session_state.raw_df if st.session_state.raw_df is not None else pd.DataFrame(),
            st.session_state.empathy_df,
            summary_df=summary
        )
        st.download_button(
            "Download combined Excel (raw + coding + summary)",
            data=excel_bytes,
            file_name=f"empathy_pipeline_{ts}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        if make_pdf and REPORTLAB_AVAILABLE:
            try:
                pdf_bytes = build_empathy_pdf(summary)
                st.download_button(
                    "Download empathy map PDF",
                    data=pdf_bytes,
                    file_name=f"empathy_maps_{ts}.pdf",
                    mime="application/pdf",
                )
            except Exception as e:
                st.error(f"Could not generate PDF: {e}")
# -------- Stage 4 --------
# -------- Stage 4 --------
with tab4:
    st.subheader("Stage 4: Draw / annotate empathy map (canvas)")

    if st.session_state.summary_df is None or st.session_state.summary_df.empty:
        st.warning("No empathy summary available yet. Please complete Stage 3 first.")
    else:
        summary = st.session_state.summary_df.copy()

        pid = st.selectbox("Select participant", summary["participant_id"].tolist())
        row = summary[summary["participant_id"] == pid].iloc[0]

        st.caption("Tip: Use the canvas tools to write notes, highlight patterns, or add sticky-note style comments.")
        st.write(f"Participant: **{pid}** | Study phase: **{row.get('study_phase','')}**")

        with st.expander("Show aggregated quotes (from Stage 3 summary)"):
            st.markdown("**THINKS & FEELS**\n" + (row.get("thinks_feels","") or "_(empty)_"))
            st.markdown("**HOW / WHERE / WHEN / TOOLS**\n" + (row.get("how_where_when_tools","") or "_(empty)_"))
            st.markdown("**MOTIVATION**\n" + (row.get("motivation","") or "_(empty)_"))
            st.markdown("**WHAT TO LEARN**\n" + (row.get("what_to_learn","") or "_(empty)_"))
            st.markdown("**PAINS**\n" + (row.get("pains","") or "_(empty)_"))
            st.markdown("**GAINS**\n" + (row.get("gains","") or "_(empty)_"))

        mode = st.selectbox(
            "Drawing tool",
            ["freedraw", "rect", "circle", "line", "transform"],
            key=f"drawing_tool_{pid}"
        )

        W, H = 1200, 850

        # Optional horizontal scroll wrapper
        st.markdown(
            """
            <div style="
                width: 100%;
                overflow-x: auto;
                overflow-y: auto;
                border: 1px solid #ddd;
                padding: 6px;
                background: #fafafa;
            ">
            """,
            unsafe_allow_html=True
        )

        canvas_result = st_canvas(
            fill_color="rgba(255, 255, 0, 0.2)",
            stroke_width=2,
            stroke_color="#000000",
            background_color="#FFFFFF",
            update_streamlit=True,
            height=H,
            width=W,
            drawing_mode=mode,
            initial_drawing=empathy_canvas_template(W, H),
            key=f"canvas_{pid}",
        )

        st.markdown("</div>", unsafe_allow_html=True)

        if canvas_result.image_data is not None:
            out = Image.fromarray(canvas_result.image_data.astype("uint8"))
            bio = BytesIO()
            out.save(bio, format="PNG")
            st.download_button(
                "Download annotated empathy map (PNG)",
                data=bio.getvalue(),
                file_name=f"empathy_map_canvas_{pid}.png",
                mime="image/png",
            )
